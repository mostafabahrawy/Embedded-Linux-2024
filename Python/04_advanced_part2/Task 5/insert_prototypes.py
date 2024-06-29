import openpyxl
import re

header_file = 'headers.h'
excel_file = 'excel.xlsx'

with open(header_file, 'r') as f:
    copy = f.read()

prototypes_pattern = r'^[a-zA-Z_][a-zA-Z0-9_]*\s+[a-zA-Z_][a-zA-Z0-9_]*\s*\([a-zA-Z0-9_\s,]*\)\s*;$'

#re.MULTILINE allows the ^ to match at the beginning of each line, not just the beginning of the string.
prototypes = re.findall(prototypes_pattern, copy, re.MULTILINE)

print(prototypes)
workbook = openpyxl.Workbook()
sheet = workbook.active 

# Add headers
sheet.cell(row=1, column=1, value='Prototype')
sheet.cell(row=1, column=2, value='ID')

row = 2
ID_counter = 1
for p in prototypes:
    sheet.cell(row=row, column=1, value=p)
    sheet.cell(row=row, column=2, value=f'IDX{ID_counter}')
    row+=1
    ID_counter+=1

workbook.save(excel_file)
